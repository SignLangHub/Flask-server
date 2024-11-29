import openpyxl
import json
import os
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed
'''
    파일의 병렬처리로 빠르게 처리
'''
# 엑셀 파일 경로 설정
file_dir = '{오리진 파일 경로 ex kor_ksl_origin_data}'
xlsx_files = glob.glob(os.path.join(file_dir, '*.xlsx'))

# JSON 파일 저장 디렉토리 설정
output_dir = '{저장할 파일 경로 ex pre_process}'
os.makedirs(output_dir, exist_ok=True)
output_file_path = os.path.join(output_dir, 'output_data.json')

# 기존 JSON 파일 불러오기 (존재하지 않으면 빈 리스트 생성)
if os.path.exists(output_file_path):
    with open(output_file_path, 'r', encoding='utf-8') as json_file:
        all_data_json = json.load(json_file)
else:
    all_data_json = []

def process_excel(file_path):
    """각 Excel 파일을 처리하여 JSON 데이터 생성"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    kor_sentence = None
    gloss_id_list = []
    start_list = []
    data_json = []

    row = 1
    print(f"Started processing file: {os.path.basename(file_path)}")
    while row <= sheet.max_row:
        if row % 100 == 0:  # 100행 단위로 진행 상황 표시
            print(f"Processing {os.path.basename(file_path)}: {row}/{sheet.max_row} rows processed.")

        row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        row_data = [val for val in row_data if val is not None]

        if not row_data:
            row += 1
            continue

        first_cell = row_data[0]

        if isinstance(first_cell, str) and first_cell == 'Korean sentence : ':
            if kor_sentence is not None:
                if start_list and gloss_id_list:
                    sorted_pairs = sorted(zip(start_list, gloss_id_list), key=lambda x: x[0])
                    sorted_start, sorted_gloss_id = zip(*sorted_pairs)
                else:
                    sorted_start, sorted_gloss_id = [], []
                ksl_value = ' '.join(map(str, sorted_gloss_id))
                data_json.append({
                    'kor': kor_sentence,
                    'ksl': ksl_value,
                    'time': list(sorted_start)
                })
            kor_sentence = row_data[1] if len(row_data) > 1 else None
            start_list = []
            gloss_id_list = []
            row += 1

        elif isinstance(first_cell, str) and first_cell.startswith('sign_gestures_'):
            for value in row_data[1:]:
                if value not in [None, '', 'gloss_id : ']:
                    gloss_id_list.append(value)
            row += 1
            next_row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            next_row_data = [val for val in next_row_data if val not in [None, '', 'start(s) : ']]
            start_list.extend(next_row_data)
            row += 1
        else:
            row += 1

    if kor_sentence is not None:
        if start_list and gloss_id_list:
            sorted_pairs = sorted(zip(start_list, gloss_id_list), key=lambda x: x[0])
            sorted_start, sorted_gloss_id = zip(*sorted_pairs)
        else:
            sorted_start, sorted_gloss_id = [], []
        ksl_value = ' '.join(map(str, sorted_gloss_id))
        data_json.append({
            'kor': kor_sentence,
            'ksl': ksl_value,
            'time': list(sorted_start)
        })

    print(f"Finished processing file: {os.path.basename(file_path)}")
    return data_json

def safe_write_to_file(filepath, data):
    """JSON 파일을 안전하게 저장"""
    try:
        with open(filepath, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"Error saving to file: {filepath}. Error: {e}")

# 병렬 처리로 모든 Excel 파일 처리
print("Processing started...")
results = []

with ThreadPoolExecutor() as executor:
    future_to_file = {executor.submit(process_excel, file_path): file_path for file_path in xlsx_files}
    
    for future in as_completed(future_to_file):
        file_path = future_to_file[future]
        try:
            result = future.result()  # 각 파일의 처리 결과
            results.extend(result)
            print(f"Successfully processed: {os.path.basename(file_path)}")

            # 중간 결과 저장
            safe_write_to_file(output_file_path, results)
            print(f"Updated JSON file after processing: {os.path.basename(file_path)}")

        except Exception as e:
            print(f"Error processing {os.path.basename(file_path)}: {e}")

# 최종 저장
print("All tasks completed. Saving final results...")
safe_write_to_file(output_file_path, results)
print(f"Final JSON data saved to {output_file_path}")
